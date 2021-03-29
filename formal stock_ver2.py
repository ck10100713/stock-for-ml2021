#!/usr/bin/env python
# coding: utf-8

# In[18]:


import json
import time 
import requests
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, colors, Alignment
import gdown


# In[2]:


def timeStamp(timeNum): 
    timeArray = time.localtime(timeNum) 
    otherStyleTime = time.strftime("%Y-%m-%d", timeArray) 
    return otherStyleTime


# In[23]:


url1 = "https://drive.google.com/u/0/uc?id=1M-6VKzAdiE6VVLRhLYgyyPFCIJ7dxa0q&export=download"
output = "stock_price.xlsx"
gdown.download(url1, output)


# In[12]:


def avg5(stock_id):
    sum=0
    wb =load_workbook('stock_price.xlsx')
    sheet = wb[stock_id]
    sheet.cell(row=1, column=7).value='MA5'
    for i in range(2,len(list(sheet.rows))):
        if i>6:
            sheet.cell(row=i, column=7).value=sum/5
            sum=sum-sheet.cell(row=i-5, column=5).value
        sum=sum+sheet.cell(row=i, column=5).value
    wb.save(filename = 'stock_price.xlsx')


# In[13]:


def avg10(stock_id):
    sum=0
    wb =load_workbook('stock_price.xlsx')
    sheet = wb[stock_id]
    sheet.cell(row=1, column=8).value='MA10'
    for i in range(2,len(list(sheet.rows))):
        if i>11:
            sheet.cell(row=i, column=8).value=sum/10
            sum=sum-sheet.cell(row=i-10, column=5).value
        sum=sum+sheet.cell(row=i, column=5).value
    wb.save(filename = 'stock_price.xlsx')


# In[14]:


def avg20(stock_id):
    sum=0
    wb =load_workbook('stock_price.xlsx')
    sheet = wb[stock_id]
    sheet.cell(row=1, column=9).value='MA20'
    for i in range(2,len(list(sheet.rows))):
        if i>21:
            sheet.cell(row=i, column=9).value=sum/20
            sum=sum-sheet.cell(row=i-20, column=5).value
        sum=sum+sheet.cell(row=i, column=5).value
    wb.save(filename = 'stock_price.xlsx')


# In[15]:


def avg240(stock_id):
    sum=0
    wb =load_workbook('stock_price.xlsx')
    sheet = wb[stock_id]
    sheet.cell(row=1, column=10).value='MA240'
    for i in range(2,len(list(sheet.rows))):
        if i>241:
            sheet.cell(row=i, column=10).value=sum/240
            sum=sum-sheet.cell(row=i-240, column=5).value
        sum=sum+sheet.cell(row=i, column=5).value
    wb.save(filename = 'stock_price.xlsx')


# In[24]:


def crawl_price():

    print("歡迎使用股價表單製作程式")
    stock_id=input('請輸入股票代碼，台股代號後方請加上.TW，ex:2330.TW\n')
    #stock_id='2330.TW'
    wb =load_workbook('stock_price.xlsx')
    
    ws = wb.create_sheet(title=stock_id)
    dest_filename = 'stock_price.xlsx'
    ws.title = stock_id
    d = datetime.now()
    url = "https://query1.finance.yahoo.com/v8/finance/chart/"+stock_id+"?period1=0&period2="+str(int(d.timestamp()))+"&interval=1d&events=history&=hP2rOschxO0"
    res = requests.get(url)
    data = json.loads(res.text)
          
    tableTitle = ['day', 'volume', 'open', 'high','close','low']
    for i in range(len(tableTitle)):
        c = i + 1
        ws.cell(row=1, column=c).value = tableTitle[i]
    for row in range(1,len(data['chart']['result'][0]['timestamp'])):
        ws.append([timeStamp(data['chart']['result'][0]['timestamp'][row])])
    for i in range(len(data['chart']['result'][0]['indicators']['quote'][0]['volume'])-1):
        c = i+1
        ws.cell(row=c+1, column=2).value = data['chart']['result'][0]['indicators']['quote'][0]['volume'][c]
        ws.cell(row=c+1, column=3).value = data['chart']['result'][0]['indicators']['quote'][0]['open'][c]
        ws.cell(row=c+1, column=4).value = data['chart']['result'][0]['indicators']['quote'][0]['high'][c]
        ws.cell(row=c+1, column=5).value = data['chart']['result'][0]['indicators']['quote'][0]['close'][c]
        ws.cell(row=c+1, column=6).value = data['chart']['result'][0]['indicators']['quote'][0]['low'][c]
        
        #把國定假日沒開市的情況清空
        if data['chart']['result'][0]['indicators']['quote'][0]['volume'][c] is None:
            ws.cell(row=c+1, column=2).value=0
    sheet = wb[stock_id]
    for i in range(2,len(list(sheet.rows))):
        while sheet.cell(row=i, column=2).value==0:
            sheet.delete_rows(i)
    wb.save(filename = dest_filename)
    
    #原本要做MA只是這個資料結構不好你可以跑跑看這個函數XD
    #avg5(stock_id)
    #avg10(stock_id)
    #avg20(stock_id)
    #avg240(stock_id)


# In[ ]:





# In[26]:


crawl_price()


# In[ ]:





# In[ ]:




